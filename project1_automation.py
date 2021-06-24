import openpyxl as xl
from openpyxl.chart import BarChart,Reference
def process_workbook(filename):
    wb = xl.load_workbook(filename) # transactions.xlsx
    sheet=wb['Sheet1']
    # cell=sheet['a1'] how to access specific cell
    # cell=sheet.cell(1,1)
    #print(cell.value)
    n=sheet.max_row

    for row in range(2,n+1):
        cell=sheet.cell(row,3)
        crt_price=cell.value * 0.9
        crt_price_cell=sheet.cell(row,4)
        crt_price_cell.value=crt_price

    values=Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    wb.save(filename)